from sqlite3 import connect
from json import load
from openpyxl import load_workbook
from hashlib import sha512
from datetime import datetime
from os import sep, listdir

def ObliczSkrot(dt, nip, nrb, iter):
    klucz = datetime.strftime(dt,'%Y%m%d') + nip + nrb
    for _ in range(iter):
        klucz = sha512(klucz.encode('UTF-8')).hexdigest()
    return klucz

def UtworzSkrotNRB(nrb, maska):
    skrot = ''
    for i in range(len(maska)):
        if maska[i] == 'X':
            skrot += 'X'
        else:
            skrot += nrb[i]
    return skrot

def WeryfikujPodatnikCzynny(dt, klucz):
    cursor.execute(f"SELECT Klucz FROM SkrotyPodatnikowCzynnych WHERE Data = '{dt}' AND Klucz = '{klucz}'")
    if cursor.fetchone() != None:
        return True
    return False

def WeryfikujPodatnikZwolniony(dt, klucz):
    cursor.execute(f"SELECT Klucz FROM SkrotyPodatnikowZwolnionych WHERE Data = '{dt}' AND Klucz = '{klucz}'")
    if cursor.fetchone() != None:
        return True
    return False

def WeryfikujPodatnik(dt, nip, nrb):
    cursor.execute(f"SELECT Liczba_transformacji FROM Naglowek WHERE Data = '{dt}'")
    iter = cursor.fetchone()
    if iter == None:
        return f'Brak danych w bazie na dzień {dt}'
    else:
        iter = iter[0]
    if WeryfikujPodatnikCzynny(dt, ObliczSkrot(dt, nip, nrb, iter)):
        return 'Rachunek podatnika znaleziony w wykazie podatników VAT czynnych'
    if WeryfikujPodatnikZwolniony(dt, ObliczSkrot(dt, nip, nrb, iter)):
        return 'Rachunek podatnika znaleziony w wykazie podatników VAT zwolnionych'
    cursor.execute(f"SELECT Maska FROM Maski WHERE Data = '{dt}' AND substr(Maska, 3, 8) = '{nrb[2:10]}'")
    for i in cursor.fetchall():
        if WeryfikujPodatnikCzynny(dt, ObliczSkrot(dt, nip, UtworzSkrotNRB(nrb, i[0]), iter)):
            return 'Rachunek wirtualny podatnika znaleziony w wykazie podatników VAT czynnych'
        if WeryfikujPodatnikZwolniony(dt, ObliczSkrot(dt, nip, UtworzSkrotNRB(nrb, i[0]), iter)):
            return 'Rachunek wirtualny podatnika znaleziony w wykazie podatników VAT zwolnionych'
    return 'Rachunek podatnika nie znaleziony w wykazie podatników VAT'

def UtworzBazeDanych():
    cursor.execute("CREATE TABLE IF NOT EXISTS Naglowek (Data DATE PRIMARY KEY ASC, Liczba_transformacji INTEGER)")
    cursor.execute("CREATE TABLE IF NOT EXISTS SkrotyPodatnikowCzynnych (Data DATE, Klucz TEXT, FOREIGN KEY(Data) REFERENCES Naglowek(Data))")
    cursor.execute("CREATE TABLE IF NOT EXISTS SkrotyPodatnikowZwolnionych (Data DATE, Klucz TEXT, FOREIGN KEY(Data) REFERENCES Naglowek(Data))")
    cursor.execute("CREATE TABLE IF NOT EXISTS Maski (Data DATE, Maska TEXT, FOREIGN KEY(Data) REFERENCES Naglowek(Data))")
    cursor.execute("CREATE UNIQUE INDEX IF NOT EXISTS indeks_czynni on SkrotyPodatnikowCzynnych (Data, Klucz)")
    cursor.execute("CREATE UNIQUE INDEX IF NOT EXISTS indeks_zwolnieni on SkrotyPodatnikowZwolnionych (Data, Klucz)")
    cursor.execute("CREATE UNIQUE INDEX IF NOT EXISTS indeks_maski on Maski (Data, Maska)")

def ObliczSumeKontrolna(plik_json):
    with open(f'.{sep}pliki_json{sep}{plik_json}','rb') as f:
        dane = f.read()
        return sha512(dane).hexdigest()

def WczytajPlikPlaski(dt):
    cursor.execute(f"SELECT Data FROM Naglowek WHERE Data = '{dt}'")
    if cursor.fetchone() == None:
        plik_json = f"{dt.strftime('%Y%m%d')}.json"
        path = f'.{sep}pliki_json{sep}{plik_json}'
        with open(f'.{sep}pliki_json{sep}{plik_json}.sha512sum') as f:
            wpis = f.readline()
            if (wpis[-14:-1] == plik_json) and (wpis[:128] == ObliczSumeKontrolna(plik_json)):
                print(f'Plik {plik_json} jest poprawny. Wczytuję plik z dnia: {dt}, proszę czekać...')
                with open(path):
                    dane = load(open(path))
                cursor.execute(f"INSERT INTO Naglowek VALUES ('{dt}', {dane['naglowek']['liczbaTransformacji']})")     
                cursor.executemany("INSERT INTO SkrotyPodatnikowCzynnych VALUES (?, ?)", [(dt, dane['skrotyPodatnikowCzynnych'][i]) for i in range(len(dane['skrotyPodatnikowCzynnych']))])
                cursor.executemany("INSERT INTO SkrotyPodatnikowZwolnionych VALUES (?, ?)", [(dt, dane['skrotyPodatnikowZwolnionych'][i]) for i in range(len(dane['skrotyPodatnikowZwolnionych']))])
                cursor.executemany("INSERT INTO Maski VALUES (?, ?)", [(dt, dane['maski'][i]) for i in range(len(dane['maski']))])
            else:
                print(f'Niezgodna suma kontrolna w pliku {plik_json}.sha512sum, plik {plik_json} nie zostanie zaczytany')
    else:
        print(f'Plik z dnia {dt} został już zaczytany')


if __name__ == '__main__':

    db = connect('biala_lista.db')
    cursor = db.cursor()
    UtworzBazeDanych()
    for plik in listdir(f'.{sep}pliki_json'):
        if len(plik) == 13:
            if plik[-5:] == '.json':
                data = datetime.strptime(plik[:8],'%Y%m%d').date()
                WczytajPlikPlaski(data)
    db.commit()

    path = 'biala_lista_test.xlsx'
    try:
        skoroszyt = load_workbook(path)
        arkusz = skoroszyt.active
        print(f'Aktualizacja statusów w pliku: {path} ...')
        for wiersz in arkusz.iter_rows(min_row=2):
            try:    
                data = wiersz[0].value.date()
                nip = wiersz[1].value
                nrb = wiersz[2].value
                wiersz[3].value = WeryfikujPodatnik(data, str(nip), str(nrb))
            except:
                wiersz[3].value = 'Nieprawidłowy format danych wejściowych'
        skoroszyt.save(path)
        skoroszyt.close()
    except:
        print(f'Brak prawidłowego pliku z danymi wejściowymi: {path}')

    db.close()
